import json
import os
import google.generativeai as genai
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
from openpyxl.worksheet.pagebreak import Break, RowBreak, ColBreak

API_KEY = 'AIzaSyCO7QtNysEUq5IwP8Q0myPVZwOi0U8cd8M'
if not API_KEY:
    raise ValueError("❌ Google API 키가 설정되지 않았습니다. 환경 변수 GOOGLE_API_KEY를 설정하세요.")

genai.configure(api_key=API_KEY)

# ✅ 사용 가능한 모델 확인 후 자동 선택
available_models = [m.name for m in genai.list_models()]
print("✅ 사용 가능한 모델 목록:", available_models)

if "models/gemini-2.0-pro-exp" in available_models:
    model_name = "models/gemini-2.0-pro-exp"
elif "models/gemini-1.5-pro" in available_models:
    model_name = "models/gemini-1.5-pro"
else:
    model_name = "models/gemini-1.5-flash"

print(f"✅ 선택된 모델: {model_name}")

def extract_json(response_text):
    """
    AI 응답에서 JSON 데이터만 추출하는 함수
    """
    try:
        # JSON 패턴 감지
        json_match = re.search(r'\{.*\}', response_text, re.DOTALL)
        if json_match:
            json_text = json_match.group(0)
            return json.loads(json_text)
        else:
            print("❌ AI 응답에서 JSON을 찾을 수 없습니다.")
            return None
    except json.JSONDecodeError:
        print("❌ AI 응답을 JSON으로 변환할 수 없습니다.")
        return None
    
def analyze_and_set_print_area(file_path, sheet_names, model_name):
    """
    Gemini AI의 응답(JSON 형식)을 기반으로 엑셀의 인쇄 영역 및 페이지 나누기 설정 (여러 시트 처리 가능)
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"❌ 파일을 찾을 수 없습니다: {file_path}")

    wb = load_workbook(file_path)

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"❌ 시트 '{sheet_name}'를 찾을 수 없습니다.")
            continue

        ws = wb[sheet_name]

        # 데이터 로드 및 프롬프트 생성
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        preview_data = df.head().to_string(index=False)

        full_prompt = f"""
        아래는 엑셀 데이터의 일부입니다 (시트명: {sheet_name}):
        {preview_data}

        이 데이터를 분석하여 한 페이지에 보기 좋게 출력될 영역을 JSON 형식으로 추천해주세요.
        무조건 해당 페이지에 있는 비어있지 않은 모든 셀은 프린트가 되어야 합니다.
        셀 간의 의존성과, 디자인 스타일링 등을 고려하여 한 페이지에 적절한 데이터들이 묶여 출력될 수 있게 해주세요.

        range의 개수는 상관이 없습니다
        형식:
        {{
            "pages": [
                {{"range": "A1:D20"}},
                {{"range": "I1:N30"}}
            ]
        }}
        설명 없이 JSON 형식으로만 출력해 주세요.
        """

        # AI 요청
        model = genai.GenerativeModel(model_name)
        response = model.generate_content(full_prompt)
        ai_settings = extract_json(response.text)
        print(ai_settings)
        if not ai_settings:
            continue

        # AI 응답에서 받은 페이지 영역 처리
        ranges = [page["range"] for page in ai_settings["pages"]]

        # 인쇄 영역 설정
        ws.print_area = ",".join(ranges)

        # 기존 나누기 제거 (올바른 방식으로 수정)
        ws.row_breaks = RowBreak()
        ws.col_breaks = ColBreak()

        # 페이지 나누기 추가 (각 영역별로 페이지 나누기)
        for area in ranges[:-1]:
            _, end = area.split(":")
            col_letter = re.findall(r'[A-Za-z]+', end)[0]
            col_idx = ws[col_letter + '1'].column
            ws.col_breaks.append(Break(id=col_idx))

        # 페이지 설정 (A4, 가로 맞춤)
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    # 저장
    output_filename = os.path.join(os.path.dirname(file_path), "modified_" + os.path.basename(file_path))
    wb.save(output_filename)
    print(f"\n✅ 변경된 파일 저장 완료: {output_filename}")

    return output_filename

# 사용 예시 (여러 시트 처리 가능)
file_path = "data/졸업이수학점정리.xlsx"
sheet_names_input = input("처리할 시트명을 쉼표로 구분하여 입력하세요: ")
sheet_names = [name.strip() for name in sheet_names_input.split(",")]
model_name = "models/gemini-2.0-pro-exp"  # 여기에 사용할 Gemini AI 모델명을 지정하세요.
modified_file = analyze_and_set_print_area(file_path, sheet_names, model_name)

modified_file
